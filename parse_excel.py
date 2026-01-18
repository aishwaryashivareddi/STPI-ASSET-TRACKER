import openpyxl
import json
from datetime import datetime

def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return str(value) if value else None

def parse_value(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    return str(value).strip() if str(value).strip() else None

def get_asset_type(sheet_name):
    mapping = {
        'HSDC-Equipments': 'HSDC',
        'Computers-peripherals': 'COMPUTER',
        'Electrical-Equipments': 'ELECTRICAL',
        'Office Equipments': 'OFFICE',
        'Furnitures-Fixtures-Misc': 'FURNITURE',
        'Fire-Fighting': 'FIREFIGHTING',
        'Building': 'BUILDING'
    }
    return mapping.get(sheet_name, 'OTHER')

wb = openpyxl.load_workbook('public/Fixed Assets Physical verification-Hyderabad FY 2024-25-latest.xlsx', data_only=True)
assets = []

sheets_to_parse = ['Building', 'HSDC-Equipments', 'Computers-peripherals', 'Electrical-Equipments', 
                   'Office Equipments', 'Furnitures-Fixtures-Misc', 'Fire-Fighting', 'not working  obsolute  ']

for sheet_name in sheets_to_parse:
    if sheet_name not in wb.sheetnames:
        continue
    
    ws = wb[sheet_name]
    asset_type = get_asset_type(sheet_name)
    
    # Find header row
    header_row = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if any(cell and 'Name of the Equipment' in str(cell) for cell in row):
            header_row = idx
            break
    
    if not header_row:
        continue
    
    # Get headers
    headers = [cell for cell in ws[header_row]]
    
    # Find column indices
    name_col = qty_col = po_col = supplier_col = invoice_col = value_col = None
    location_col = serial_col = barcode_col = status_col = remarks_col = None
    
    for idx, cell in enumerate(headers):
        if cell.value:
            val = str(cell.value).lower()
            if 'name of the equipment' in val or 'particulars' in val:
                name_col = idx
            elif val == 'qty' or 'quantity' in val:
                qty_col = idx
            elif 'purchase order' in val or 'po no' in val:
                po_col = idx
            elif 'supplier' in val:
                supplier_col = idx
            elif 'invoice' in val:
                invoice_col = idx
            elif 'purchase value' in val or 'value' in val:
                value_col = idx
            elif 'location' in val:
                location_col = idx
            elif 'sl.no' in val or 'serial' in val:
                serial_col = idx
            elif 'barcode' in val or 'ams' in val:
                barcode_col = idx
            elif 'status' in val:
                status_col = idx
            elif 'remark' in val:
                remarks_col = idx
    
    # Parse data rows
    for row in ws.iter_rows(min_row=header_row+2, values_only=False):
        name = row[name_col].value if name_col is not None else None
        if not name or str(name).strip() in ['', 'Total', 'TOTAL']:
            continue
        
        qty = row[qty_col].value if qty_col is not None else 1
        # Validate quantity - should be a reasonable number
        if isinstance(qty, (int, float)) and qty > 0 and qty < 10000:
            qty = int(qty)
        else:
            qty = 1
        
        asset = {
            'asset_type': asset_type,
            'name': parse_value(name),
            'quantity': qty,
            'branch_id': 1,
            'location': parse_value(row[location_col].value) if location_col is not None else None,
            'serial_number': parse_value(row[serial_col].value) if serial_col is not None else None,
            'ams_barcode': parse_value(row[barcode_col].value) if barcode_col is not None else None,
            'supplier_name': parse_value(row[supplier_col].value) if supplier_col is not None else None,
            'po_number': parse_value(row[po_col].value) if po_col is not None else None,
            'po_date': parse_date(row[po_col].value) if po_col is not None else None,
            'invoice_date': parse_date(row[invoice_col].value) if invoice_col is not None else None,
            'purchase_value': row[value_col].value if value_col is not None and isinstance(row[value_col].value, (int, float)) else None,
            'current_status': parse_value(row[status_col].value) if status_col is not None else 'Working',
            'remarks': parse_value(row[remarks_col].value) if remarks_col is not None else None
        }
        assets.append(asset)

with open('assets_seed_data.json', 'w', encoding='utf-8') as f:
    json.dump(assets, f, indent=2, ensure_ascii=False)

total_qty = sum(a['quantity'] for a in assets)
print(f'Parsed {len(assets)} asset entries')
print(f'Total quantity: {total_qty}')
print(f'\nBreakdown by type:')
by_type = {}
for a in assets:
    by_type[a['asset_type']] = by_type.get(a['asset_type'], 0) + a['quantity']
for k, v in sorted(by_type.items()):
    print(f'  {k}: {v}')
